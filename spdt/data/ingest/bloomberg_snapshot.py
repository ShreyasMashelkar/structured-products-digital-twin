"""Bloomberg frozen-snapshot source — a fourth :class:`MarketDataSource` (L1).

A one-time Bloomberg Terminal export (see ``tools/bloomberg/build_pull_template.py``) is a
*static* pull, not a live feed: everything is captured as of a single business date and saved
to a workbook. This source reads that workbook back and emits the same immutable
:class:`RawMarketData` the synthetic/NSE/Dhan sources do, so nothing downstream knows the
difference — the snapshot-in / report-out invariant is preserved.

Bloomberg supplies **implied vols** (from ``OVDV`` or an option chain), not settlement prices,
so we convert each grid point IV → Black-76 price via :func:`spdt.data.curate.bs_price`. The
downstream BS inversion then recovers exactly the Bloomberg smile, which is what makes the
whole ingestion → snapshot path testable against a known surface — identical in spirit to the
synthetic source, but the numbers are real and tagged :attr:`SourceTag.OBSERVED`.

Expected workbook layout (produced by the template generator, filled on the Terminal):

* sheet ``Scalars`` — a ``key | value`` table with ``spot``, ``dividend_yield`` and an
  optional ``funding_spread`` (decimals; percents auto-detected).
* sheet ``OIS`` — ``tenor_years | zero_rate`` rows (continuously-compounded).
* sheet ``VolSurface`` — a grid: row 1 (from col B) tenors in years, col A (from row 2)
  moneyness ``K/S``, interior cells implied vol.

Only the equity/vol/rates blocks feed :class:`RawMarketData`. The credit (CDS/``DRSK``) and
dividend-schedule blocks in the same workbook are consumed separately by the XVA seam
(``integration/credit.py``) and the dividend curve builder; they are intentionally not folded
in here so this source's output shape stays identical to the other three.
"""

from __future__ import annotations

from datetime import date as Date
from datetime import timedelta
from math import exp
from pathlib import Path

from spdt.core.types import SourceTag, Underlying, year_fraction
from spdt.data.curate.bs_inversion import bs_price
from spdt.data.ingest import RawMarketData, RawOptionQuote

_DEFAULT_FUNDING_SPREAD = 0.012


def _to_decimal(value: float, *, percent_threshold: float) -> float:
    """Normalise a rate/vol that may have been exported in percent (e.g. ``18`` vs ``0.18``)."""
    return value / 100.0 if abs(value) > percent_threshold else value


def _interp(xs: list[float], ys: list[float], x: float) -> float:
    """Linear interpolation over sorted ``xs`` with flat extrapolation at both ends."""
    if x <= xs[0]:
        return ys[0]
    if x >= xs[-1]:
        return ys[-1]
    for i in range(1, len(xs)):
        if x <= xs[i]:
            w = (x - xs[i - 1]) / (xs[i] - xs[i - 1])
            return ys[i - 1] + w * (ys[i] - ys[i - 1])
    return ys[-1]


class BloombergSnapshotSource:
    """Read a filled Bloomberg-export workbook into :class:`RawMarketData` for one date.

    Parameters
    ----------
    workbook_path:
        Path to the ``.xlsx`` exported and saved on the Terminal (cached values are read; the
        workbook need not recalculate off-Terminal).
    funding_spread:
        Fallback issuer funding spread over OIS if the workbook omits it (ADR-0002).
    """

    def __init__(
        self, workbook_path: str | Path, *, funding_spread: float = _DEFAULT_FUNDING_SPREAD
    ) -> None:
        self._path = Path(workbook_path)
        self._funding_spread = funding_spread

    def fetch(self, as_of: Date, underlying: Underlying = "NIFTY") -> RawMarketData:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover - optional dep
            raise RuntimeError(
                "reading a Bloomberg workbook needs openpyxl: pip install openpyxl"
            ) from exc

        wb = load_workbook(self._path, data_only=True, read_only=True)
        try:
            scalars = self._read_scalars(wb["Scalars"])
            tenors, zeros = self._read_ois(wb["OIS"])
            moneyness, grid_tenors, vols = self._read_surface(wb["VolSurface"])
        finally:
            wb.close()

        spot = scalars["spot"]
        div_yield = _to_decimal(scalars["dividend_yield"], percent_threshold=0.5)
        funding = _to_decimal(
            scalars.get("funding_spread", self._funding_spread), percent_threshold=0.5
        )

        # Rate pillars keyed by calendar date, mirroring the synthetic source's shape.
        ois_zero_rates = {
            as_of + timedelta(days=round(t * 365)): _to_decimal(z, percent_threshold=1.0)
            for t, z in zip(tenors, zeros, strict=True)
        }
        pillar_dates = sorted(ois_zero_rates)
        funding_spread_knots = {pillar_dates[0]: funding, pillar_dates[-1]: funding}

        rate_at = lambda t: _interp(  # noqa: E731 - tiny local closure over the read curve
            tenors, [_to_decimal(z, percent_threshold=1.0) for z in zeros], t
        )

        quotes: list[RawOptionQuote] = []
        for j, tenor in enumerate(grid_tenors):
            expiry = as_of + timedelta(days=round(tenor * 365))
            tau = year_fraction(as_of, expiry)
            if tau <= 0.0:
                continue
            rate = rate_at(tenor)
            forward = spot * exp((rate - div_yield) * tau)
            discount = exp(-rate * tau)
            for i, m in enumerate(moneyness):
                sigma = vols[i][j]
                if sigma is None:
                    continue
                sigma = _to_decimal(float(sigma), percent_threshold=1.5)
                strike = spot * m
                for is_call in (True, False):
                    price = bs_price(forward, strike, tau, sigma, discount, is_call)
                    quotes.append(RawOptionQuote(expiry, strike, is_call, price))

        return RawMarketData(
            date=as_of,
            underlying=underlying,
            spot=spot,
            option_chain=tuple(quotes),
            ois_zero_rates=ois_zero_rates,
            funding_spread_knots=funding_spread_knots,
            dividend_yield=div_yield,
            source=SourceTag.OBSERVED,
        )

    @staticmethod
    def _read_scalars(ws) -> dict[str, float]:  # type: ignore[no-untyped-def]
        out: dict[str, float] = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or row[0] is None or len(row) < 2 or row[1] is None:
                continue
            key = str(row[0]).strip().lower().replace(" ", "_")
            try:
                out[key] = float(row[1])
            except (TypeError, ValueError):
                continue
        missing = {"spot", "dividend_yield"} - out.keys()
        if missing:
            raise ValueError(f"Scalars sheet missing required keys: {sorted(missing)}")
        return out

    @staticmethod
    def _read_ois(ws) -> tuple[list[float], list[float]]:  # type: ignore[no-untyped-def]
        tenors: list[float] = []
        zeros: list[float] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or len(row) < 2 or row[1] is None:
                continue
            try:
                tenors.append(float(row[0]))
                zeros.append(float(row[1]))
            except (TypeError, ValueError):
                continue
        if len(tenors) < 2:
            raise ValueError("OIS sheet needs at least two tenor/rate rows")
        order = sorted(range(len(tenors)), key=lambda i: tenors[i])
        return [tenors[i] for i in order], [zeros[i] for i in order]

    @staticmethod
    def _read_surface(  # type: ignore[no-untyped-def]
        ws,
    ) -> tuple[list[float], list[float], list[list[float | None]]]:
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            raise ValueError("VolSurface sheet is empty")
        header = rows[0]
        grid_tenors = [float(v) for v in header[1:] if v is not None]
        moneyness: list[float] = []
        vols: list[list[float | None]] = []
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                moneyness.append(float(row[0]))
            except (TypeError, ValueError):
                continue
            cells: list[float | None] = []
            for k in range(len(grid_tenors)):
                cell = row[1 + k] if 1 + k < len(row) else None
                cells.append(float(cell) if cell is not None else None)
            vols.append(cells)
        if not moneyness or not grid_tenors:
            raise ValueError("VolSurface needs a tenor header row and moneyness column")
        return moneyness, grid_tenors, vols


__all__ = ["BloombergSnapshotSource"]
