"""Bloomberg rates overlay for the user's Terminal export.

The workbook ``Data for Intern's usage.xlsx`` is not an equity option-chain export.  It contains
three market blocks on one sheet:

* Modified MIFOR curve / USDINR forwards
* SOFR curve
* USDINR volatility surface

Only the MIFOR block can defensibly influence SPDT's INR funding/projection assumptions for a
NIFTY structured-products run.  It should **not** replace the OIS/MIBOR-style INR discount curve:
MIFOR embeds USD rates and USD/INR forward/basis effects.  The USDINR volatility surface is market
data, but it is FX vol, not NIFTY equity vol, so this module deliberately does **not** feed it into
the equity smile.

The class below wraps another :class:`MarketDataSource` (synthetic/NSE/Dhan), leaves its equity
inputs and OIS/MIBOR-style discount curve untouched, and replaces only issuer funding-spread knots
with the spread implied by Bloomberg MIFOR over the base INR discount curve.
"""

from __future__ import annotations

from dataclasses import replace
from datetime import date as Date
from datetime import timedelta
from pathlib import Path
from typing import Any

from spdt.core.types import Underlying
from spdt.data.ingest import MarketDataSource, RawMarketData


def _excel_yyyymmdd(value: Any) -> Date | None:
    try:
        text = str(int(value))
    except (TypeError, ValueError):
        return None
    if len(text) != 8:
        return None
    try:
        return Date(int(text[:4]), int(text[4:6]), int(text[6:8]))
    except ValueError:
        return None


def _rate_decimal(value: Any) -> float | None:
    try:
        x = float(value)
    except (TypeError, ValueError):
        return None
    # Bloomberg pages commonly export rates as percent points.
    return x / 100.0 if abs(x) > 1.0 else x


class BloombergRatesOverlaySource:
    """Overlay Bloomberg-exported MIFOR funding spreads onto an existing market-data source.

    Parameters
    ----------
    base:
        Source providing the equity inputs.  For offline runs this is normally
        :class:`SyntheticSource`.
    workbook_path:
        One-sheet Bloomberg workbook with a ``Modified Mifor curve`` block in columns A:D.
    funding_spread:
        Fallback spread if a MIFOR pillar cannot be mapped.  The workbook does not contain issuer
        credit/funding spreads, so this is a floor/model assumption, not a credit curve.
    """

    def __init__(
        self,
        base: MarketDataSource,
        workbook_path: str | Path,
        *,
        funding_spread: float = 0.012,
    ) -> None:
        self._base = base
        self._path = Path(workbook_path)
        self._funding_spread = funding_spread

    def fetch(self, as_of: Date, underlying: Underlying = "NIFTY") -> RawMarketData:
        raw = self._base.fetch(as_of, underlying)
        curve = self._read_mifor_curve(as_of)
        if len(curve) < 2:
            raise ValueError(
                f"Bloomberg workbook {self._path} did not contain enough usable MIFOR rate pillars"
            )
        funding_spread_knots = self._mifor_spread_over_base_ois(raw, curve)
        return replace(
            raw,
            funding_spread_knots=funding_spread_knots,
            # Keep the base source tag. The snapshot is mixed, and RawMarketData has one coarse
            # source field; field-level honesty is exposed through the dashboard data_boundary.
            source=raw.source,
        )

    def _read_mifor_curve(self, as_of: Date) -> dict[Date, float]:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError("reading the Bloomberg workbook needs openpyxl") from exc

        wb = load_workbook(self._path, data_only=True, read_only=True)
        try:
            ws = wb.active
            pillars: dict[Date, float] = {}
            for row in ws.iter_rows(min_row=4, values_only=True):
                if len(row) < 4:
                    continue
                term, unit, rate, rate_type = row[0], row[1], row[2], row[3]
                if str(rate_type or "").strip().lower() != "swap":
                    continue
                if str(unit or "").strip().upper() != "YR":
                    continue
                try:
                    tenor_years = float(term)
                except (TypeError, ValueError):
                    continue
                dec = _rate_decimal(rate)
                if dec is None:
                    continue
                pillars[as_of + timedelta(days=round(365 * tenor_years))] = dec
        finally:
            wb.close()

        # Add a short anchor by flat-extending the first available swap rate.  The export's
        # sub-1Y rows are FX forward points, not interest rates, so using them as rates would be
        # worse than a transparent short-end extrapolation.
        if pillars:
            first_rate = pillars[min(pillars)]
            for days in (30, 91, 182, 365):
                pillars.setdefault(as_of + timedelta(days=days), first_rate)
        return dict(sorted(pillars.items()))

    @staticmethod
    def _rate_for(expiry: Date, curve: dict[Date, float]) -> float:
        if expiry in curve:
            return curve[expiry]
        before = [d for d in curve if d <= expiry]
        after = [d for d in curve if d >= expiry]
        if not before:
            return curve[min(curve)]
        if not after:
            return curve[max(curve)]
        lo, hi = max(before), min(after)
        if lo == hi:
            return curve[lo]
        t0, t1 = lo.toordinal(), hi.toordinal()
        w = (expiry.toordinal() - t0) / (t1 - t0)
        return curve[lo] + w * (curve[hi] - curve[lo])

    def _mifor_spread_over_base_ois(
        self, raw: RawMarketData, mifor_curve: dict[Date, float]
    ) -> dict[Date, float]:
        """Map the Bloomberg MIFOR curve to issuer funding-spread knots over base OIS.

        SPDT's funding curve is represented as ``base OIS + spread``.  Because the workbook lacks
        MIBOR/OIS, the base discount curve must remain the wrapped source's curve; MIFOR contributes
        only an indicative funding/projection spread.
        """
        spreads: dict[Date, float] = {}
        for pillar, mifor_rate in mifor_curve.items():
            base_rate = self._rate_for(pillar, dict(raw.ois_zero_rates))
            spreads[pillar] = max(mifor_rate - base_rate, 0.0)
        if not spreads:
            first, last = min(raw.ois_zero_rates), max(raw.ois_zero_rates)
            return {first: self._funding_spread, last: self._funding_spread}
        return dict(sorted(spreads.items()))


__all__ = ["BloombergRatesOverlaySource"]
