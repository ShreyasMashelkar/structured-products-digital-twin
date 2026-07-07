"""Generate the Bloomberg one-shot pull workbook (and a testable filled example).

Run on your machine; the *template* is what your friend opens on the college Terminal. It
carries live ``BDP``/``BDH``/``BDS`` formulas plus paste targets for the screens that only
export as grids (``OVDV`` vol surface, ``ICVS`` OIS curve, ``CDSW``/``DRSK`` credit). He opens
it, lets it populate, and ``Save As`` → emails it back. Reading it back into SPDT is
:class:`spdt.data.ingest.bloomberg_snapshot.BloombergSnapshotSource`.

Because the Excel add-in has a **monthly data budget**, this is sized for one focused session:
a coherent cross-section as of one date plus a modest price history for correlation. Don't
have him explore at the Terminal — this file is the whole shopping list.

    python tools/bloomberg/build_pull_template.py

writes ``tools/bloomberg/bloomberg_pull_template.xlsx`` (for the Terminal) and
``tools/bloomberg/bloomberg_example_filled.xlsx`` (synthetic numbers, for offline loader tests).
"""

from __future__ import annotations

from math import log
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── what to pull ────────────────────────────────────────────────────────────────────────
INDEX_TICKER = "NIFTY Index"
# ~10 liquid NSE names with listed options and real dividends. Edit to your product universe.
SINGLE_NAMES = [
    "RELIANCE IN Equity", "HDFCBANK IN Equity", "ICICIBANK IN Equity",
    "INFY IN Equity", "TCS IN Equity", "ITC IN Equity",
    "LT IN Equity", "SBIN IN Equity", "AXISBANK IN Equity", "BHARTIARTL IN Equity",
]
# Surface grid the loader expects: moneyness (K/S) down col A, tenor-years across row 1.
MONEYNESS = [0.80, 0.90, 0.95, 1.00, 1.05, 1.10, 1.20]
TENORS_Y = [0.25, 0.50, 1.00, 2.00, 3.00, 5.00]
OIS_TENORS_Y = [0.08, 0.25, 0.50, 1.00, 2.00, 3.00, 5.00]
HISTORY_START = "-5CY"  # BDH: 5 calendar years back (Bloomberg relative date)

_HDR = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_NOTE = Font(italic=True, color="555555")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _header(ws, cells: list[str], row: int = 1) -> None:
    for c, text in enumerate(cells, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = _HDR
        cell.fill = _HDR_FILL


def _instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    steps = [
        ("SPDT Bloomberg pull — one session, one snapshot", True),
        ("", False),
        ("This file fills ITSELF. Your only job here is: open it on the Bloomberg PC, wait for "
         "the blue cells to turn into numbers, then Save As and send it back.", False),
        ("", False),
        ("1. Open on the Bloomberg PC with the Excel add-in installed. The 'Scalars', 'Spot', "
         "'History' and 'Dividends' tabs fill automatically once the add-in connects (~30s) — "
         "just wait for every #N/A to resolve to a number.", False),
        ("2. File -> Save As -> keep .xlsx -> name it template_filled_<today> -> send it back.", False),
        ("", False),
        ("The vol surface, OIS curve and credit data are NOT in this file. You export those from "
         "their Bloomberg screens as SEPARATE Excel files — see FRIEND_GUIDE for the exact steps. "
         "(This keeps things simple: no pasting into this workbook.)", False),
        ("", False),
        ("Data-budget note: the add-in caps monthly downloads. Don't add extra tickers or long "
         "histories — this list is sized to fit. If a cell shows a limit error, tell Harsh.", False),
    ]
    for i, (text, is_title) in enumerate(steps, start=1):
        cell = ws.cell(row=i, column=2, value=text)
        cell.alignment = _WRAP
        cell.font = Font(bold=True, size=14) if is_title else (_NOTE if text.startswith(("Data", "This")) else Font())


def _scalars(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("Scalars")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    _header(ws, ["key", "value"])
    rows = [
        ("underlying", "NIFTY", "NIFTY"),
        ("spot", 24100.0, f'=BDP("{INDEX_TICKER}","PX_LAST")'),
        ("dividend_yield", 1.30, f'=BDP("{INDEX_TICKER}","EQY_DVD_YLD_IND")'),  # percent; auto-normalised
        ("funding_spread", None, None),  # optional; loader falls back to its default
    ]
    for r, (key, demo, formula) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=demo if filled else formula)
    ws.cell(row=6, column=1, value="note").font = _NOTE
    ws.cell(row=6, column=2,
            value="spot & dividend_yield auto-fill via BDP. Leave funding_spread blank.").font = _NOTE


def _spot(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("Spot")
    ws.column_dimensions["A"].width = 26
    for col in "BCD":
        ws.column_dimensions[col].width = 14
    _header(ws, ["ticker", "px_last", "dvd_yld_ind", "px_volume"])
    tickers = [INDEX_TICKER, *SINGLE_NAMES]
    for r, tk in enumerate(tickers, start=2):
        ws.cell(row=r, column=1, value=tk)
        if filled:
            ws.cell(row=r, column=2, value=round(100 + r * 37.5, 2))
            ws.cell(row=r, column=3, value=round(0.8 + 0.1 * r, 2))
            ws.cell(row=r, column=4, value=1_000_000 + r)
        else:
            ws.cell(row=r, column=2, value=f'=BDP(A{r},"PX_LAST")')
            ws.cell(row=r, column=3, value=f'=BDP(A{r},"EQY_DVD_YLD_IND")')
            ws.cell(row=r, column=4, value=f'=BDP(A{r},"PX_VOLUME")')


def _history(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("History")
    ws.column_dimensions["A"].width = 60
    _header(ws, ["daily close history (for correlation) — one BDH block per ticker"])
    ws.cell(row=2, column=1).font = _NOTE
    ws.cell(row=2, column=1,
            value="Each BDH spills a Date|PX_LAST table below its anchor cell. "
                  "Keep the name count modest to stay under the data budget.").font = _NOTE
    tickers = [INDEX_TICKER, *SINGLE_NAMES]
    row = 4
    for tk in tickers:
        ws.cell(row=row, column=1, value=tk).font = Font(bold=True)
        if not filled:
            ws.cell(row=row + 1, column=1,
                    value=f'=BDH("{tk}","PX_LAST","{HISTORY_START}","TODAY")')
        else:
            ws.cell(row=row + 1, column=1, value="2021-07-01")
            ws.cell(row=row + 1, column=2, value=100.0)
        row += 8  # leave room for the spilled series


def _vol_surface(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("VolSurface")
    ws.column_dimensions["A"].width = 16
    ws.cell(row=1, column=1, value="moneyness \\ tenor(y)").font = _HDR
    ws.cell(row=1, column=1).fill = _HDR_FILL
    for j, t in enumerate(TENORS_Y, start=2):
        cell = ws.cell(row=1, column=j, value=t)
        cell.font = _HDR
        cell.fill = _HDR_FILL
    shade = PatternFill("solid", fgColor="FFF2CC")
    for i, m in enumerate(MONEYNESS, start=2):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        for j, t in enumerate(TENORS_Y, start=2):
            cell = ws.cell(row=i, column=j)
            if filled:
                k = log(m)  # simple skewed smile in percent, decays with tenor
                cell.value = round(100 * (0.20 - 0.05 * k + 0.10 * k * k - 0.01 * t), 2)
            else:
                cell.fill = shade  # paste OVDV implied vols here (percent or decimal both fine)
    note_row = len(MONEYNESS) + 3
    ws.cell(row=note_row, column=1,
            value="Paste the OVDV implied-vol grid into the shaded cells. Match this grid's "
                  "moneyness (col A) and tenors (row 1); percent or decimal both read fine.").font = _NOTE


def _ois(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("OIS")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    _header(ws, ["tenor_years", "zero_rate"])
    for r, t in enumerate(OIS_TENORS_Y, start=2):
        ws.cell(row=r, column=1, value=t)
        if filled:
            ws.cell(row=r, column=2, value=round(6.40 + 0.05 * t, 3))  # percent
        # else: leave blank — paste from ICVS export
    ws.cell(row=len(OIS_TENORS_Y) + 3, column=1,
            value="Paste INR OIS zero rates (from ICVS) next to each tenor. Optional cross-check "
                  "of SPDT's FBIL curve.").font = _NOTE


def _dividends(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("Dividends")
    ws.column_dimensions["A"].width = 26
    _header(ws, ["projected discrete dividends (ex-date + amount) per name"])
    ws.cell(row=2, column=1,
            value="Autocallables are dividend-sensitive — this is the discrete schedule, not "
                  "just the yield.").font = _NOTE
    row = 4
    for tk in SINGLE_NAMES:
        ws.cell(row=row, column=1, value=tk).font = Font(bold=True)
        if not filled:
            ws.cell(row=row + 1, column=1,
                    value=f'=BDS("{tk}","BDVD_PR_EX_DTS_DVD_AMTS")')
        row += 6


def _credit(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("Credit")
    for col in "ABC":
        ws.column_dimensions[col].width = 20
    _header(ws, ["tenor_years", "cds_spread_bp", "default_prob_pct"])
    for r, t in enumerate([0.5, 1, 2, 3, 5, 7, 10], start=2):
        ws.cell(row=r, column=1, value=t)
        if filled:
            ws.cell(row=r, column=2, value=round(80 + 12 * t, 1))
            ws.cell(row=r, column=3, value=round(0.5 + 0.3 * t, 2))
    ws.cell(row=11, column=1,
            value="Paste a CDS term structure (CDSW: India sovereign + a bank) and DRSK issuer "
                  "default probabilities. Feeds the XVA seam, not the equity snapshot.").font = _NOTE


def _validation(wb: Workbook, filled: bool) -> None:
    ws = wb.create_sheet("Validation")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    _header(ws, ["DLIB benchmark — price THIS note, then type the result + screenshot"])
    terms = [
        ("underlying", INDEX_TICKER),
        ("product", "autocallable (worst-of N/A, single underlier)"),
        ("notional", 1_000_000),
        ("maturity_years", 3),
        ("observation_freq", "annual"),
        ("autocall_barrier_pct", 100),
        ("coupon_barrier_pct", 70),
        ("coupon_pa_pct", "(solve / read)"),
        ("protection_barrier_pct", 65),
        ("", ""),
        ("DLIB price", "<< type here >>"),
        ("DLIB delta", "<< type here >>"),
        ("DLIB vega", "<< type here >>"),
        ("screenshot filename", "<< attach DLIB_autocall.png >>"),
    ]
    for r, (k, v) in enumerate(terms, start=2):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True) if k else Font()
        ws.cell(row=r, column=2, value=v)


def build(out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    template_path = out_dir / "bloomberg_pull_template.xlsx"
    example_path = out_dir / "bloomberg_example_filled.xlsx"
    # The friend-facing TEMPLATE holds only the self-filling (BDP/BDH/BDS) tabs — the vol
    # surface, OIS and credit arrive as their own Bloomberg exports (see FRIEND_GUIDE), so
    # they're deliberately absent here to avoid empty "am I meant to fill this?" tabs.
    # The EXAMPLE keeps every tab: it's a self-contained, synthetic all-in-one fixture that
    # exercises the IV→price→snapshot path offline (see BloombergSnapshotSource).
    for filled, path, full in ((False, template_path, False), (True, example_path, True)):
        wb = Workbook()
        _instructions(wb)
        _scalars(wb, filled)
        _spot(wb, filled)
        _history(wb, filled)
        _dividends(wb, filled)
        if full:
            _vol_surface(wb, filled)
            _ois(wb, filled)
            _credit(wb, filled)
            _validation(wb, filled)
        wb.save(path)
    return template_path, example_path


if __name__ == "__main__":
    here = Path(__file__).resolve().parent
    tmpl, ex = build(here)
    print(f"template (for the Terminal): {tmpl}")
    print(f"example (offline loader test): {ex}")
